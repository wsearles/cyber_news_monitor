#!/usr/bin/env python3
"""
cyberNewsMonitor.py

Pulls headlines from well-known cybersecurity RSS feeds, skips anything
you've already seen (tracked in a small local SQLite database), and
highlights items that look like new CVEs, data breaches, ransomware
attacks, or actively-exploited zero-days.

Usage:
    python3 cyberNewsMonitor.py                  # run once, print results
    python3 cyberNewsMonitor.py --all             # show everything, not just matches
    python3 cyberNewsMonitor.py --watch --interval 30   # poll every 30 min, forever
    python3 cyberNewsMonitor.py --digest-dir ~/cyber-digests   # also save a Markdown file
    python3 cyberNewsMonitor.py --html-path            # save an HTML report to the Desktop & open it
    python3 cyberNewsMonitor.py --html-path ~/cyber-news.html   # ...or save it to a custom path

First run note: each feed only ever returns its most recent ~20-50 items,
so you won't get flooded with years of history. The --lookback-hours flag
(default 48) additionally hides anything older than that on top of dedup.
"""

import argparse
import calendar
import hashlib
import html
import json
import os
import platform
import re
import sqlite3
import subprocess
import sys
import time
import urllib.request
import webbrowser
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

try:
    import feedparser
except ImportError:
    sys.exit("Missing dependency. Run: pip install -r cyberNewsRequirements.txt")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing dependency. Run: pip install -r cyberNewsRequirements.txt")

XLSX_HEADERS = ["Title", "Categories", "CVEs", "Source Feed", "Published (UTC)", "Link", "First Seen (UTC)"]
XLSX_COLUMN_WIDTHS = [70, 30, 22, 24, 18, 70, 18]
XLSX_DEFAULT_FILENAME = "cyber_security_news.xlsx"

HTML_HEADERS = ["Title", "Categories", "CVEs", "Source Feed", "Published (UTC)", "Link", "First Seen (UTC)"]
HTML_DEFAULT_FILENAME = "cyber_security_news.html"
HTML_DESKTOP_DEFAULT = "__desktop__"  # sentinel: --html-path passed with no value


# --------------------------------------------------------------------------
# Default feed list (well-known, currently-active cybersecurity RSS feeds).
# You can override/extend this by passing --feeds-file pointing at a JSON
# file with the same [{"name": ..., "url": ...}, ...] structure.
# Note: CISA retired its public RSS feeds for alerts/KEV in May 2025 in
# favor of email/social notifications, so it's intentionally not listed.
# --------------------------------------------------------------------------
DEFAULT_FEEDS = [
    {"name": "Krebs on Security", "url": "https://krebsonsecurity.com/feed/"},
    {"name": "BleepingComputer", "url": "https://www.bleepingcomputer.com/feed/"},
    {"name": "The Hacker News", "url": "https://feeds.feedburner.com/TheHackersNews"},
    {"name": "Dark Reading", "url": "https://www.darkreading.com/rss.xml"},
    {"name": "SecurityWeek", "url": "https://www.securityweek.com/feed/"},
    {"name": "CSO Online", "url": "https://www.csoonline.com/feed/"},
    {"name": "Graham Cluley", "url": "https://grahamcluley.com/feed/"},
    {"name": "SANS Internet Storm Center", "url": "https://isc.sans.edu/rssfeed_full.xml"},
    {"name": "Wired Security", "url": "https://www.wired.com/feed/category/security/latest/rss"},
    {"name": "The Register Security", "url": "https://www.theregister.com/security/headlines.atom"},
    {"name": "404 Media", "url": "https://www.404media.co/rss/"},
    {"name": "Zero Day (Kim Zetter)", "url": "https://www.zetter-zeroday.com/feed"},
    {"name": "Ars Technica Security", "url": "https://arstechnica.com/security/feed/"},
    # Whole Microsoft Security Blog -- the site doesn't expose a separate
    # feed scoped to just the "Threat Intelligence" topic page, but threat
    # intel is most of what they publish here anyway.
    {"name": "Microsoft Security Blog", "url": "https://www.microsoft.com/en-us/security/blog/feed/"},
    {"name": "Securelist (Kaspersky)", "url": "https://securelist.com/feed/"},
    {"name": "FortiGuard Labs Threat Research", "url": "https://feeds.fortinet.com/fortinet/blog/threat-research"},
    # Healthcare breach/incident news -- directly relevant given our sector.
    {"name": "HIPAA Journal", "url": "https://www.hipaajournal.com/feed/"},
    {"name": "The Guardian Security", "url": "https://www.theguardian.com/technology/data-computer-security/rss"},
    # OT/ICS-focused -- relevant given connected medical device exposure.
    {"name": "Industrial Cyber", "url": "https://industrialcyber.co/feed"},
    # Raw vulnerability-disclosure mailing list, not curated news -- expect
    # more technical/PoC-heavy entries than the other feeds.
    {"name": "Seclists Full Disclosure", "url": "https://seclists.org/rss/fulldisclosure.rss"},
]

# --------------------------------------------------------------------------
# Categorization rules. Kept simple and transparent on purpose -- tweak
# these lists to taste.
# --------------------------------------------------------------------------
CVE_PATTERN = re.compile(r"CVE-\d{4}-\d{4,7}", re.IGNORECASE)

RANSOMWARE_TERMS = [
    "ransomware", "ransom note", "double extortion", "decryptor",
    "lockbit", "blackcat", "alphv", "conti", "revil", "cl0p", "clop",
    "akira", "ransomhub", "blacksuit", "play ransomware", "qilin", 'shinyhubters'
]

BREACH_TERMS = [
    "data breach", "breached", "leaked database", "exposed database",
    "unauthorized access", "compromised accounts", "stolen data",
    "exposed records", "data leak", "leaked data", "hacked and leaked",
    "customer data exposed",
]

ZERO_DAY_TERMS = [
    "zero-day", "0-day", "actively exploited", "exploited in the wild",
    "proof-of-concept exploit", "poc exploit", "under active exploitation",
]

CATEGORY_ORDER = ["CVE", "Zero-Day/Actively Exploited", "Ransomware", "Breach"]


@dataclass
class Item:
    feed: str
    title: str
    link: str
    summary: str
    published: Optional[datetime]
    categories: list = field(default_factory=list)
    cves: list = field(default_factory=list)

    def sort_key(self):
        return self.published or datetime.fromtimestamp(0, tz=timezone.utc)


def entry_id(entry) -> str:
    """Stable unique id for dedup, preferring the feed's own guid."""
    raw = entry.get("id") or entry.get("link") or entry.get("title", "")
    return hashlib.sha256(raw.encode("utf-8", "ignore")).hexdigest()


def parse_published(entry) -> Optional[datetime]:
    for field_name in ("published_parsed", "updated_parsed"):
        struct = entry.get(field_name)
        if struct:
            return datetime.fromtimestamp(calendar.timegm(struct), tz=timezone.utc)
    return None


def categorize(title: str, summary: str):
    text = f"{title}\n{summary}".lower()
    categories = []
    cves = sorted(set(CVE_PATTERN.findall(f"{title}\n{summary}")), key=str.upper)
    if cves:
        categories.append("CVE")
    if any(term in text for term in ZERO_DAY_TERMS):
        categories.append("Zero-Day/Actively Exploited")
    if any(term in text for term in RANSOMWARE_TERMS):
        categories.append("Ransomware")
    if any(term in text for term in BREACH_TERMS):
        categories.append("Breach")
    return categories, cves


# --------------------------------------------------------------------------
# Local "seen" store -- a tiny SQLite db so re-runs don't repeat themselves.
# --------------------------------------------------------------------------
def open_db(path: Path) -> sqlite3.Connection:
    path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(path)
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS seen (
            id TEXT PRIMARY KEY,
            feed TEXT,
            title TEXT,
            link TEXT,
            published TEXT,
            first_seen TEXT
        )
        """
    )
    conn.commit()
    return conn


def has_seen(conn: sqlite3.Connection, item_id: str) -> bool:
    cur = conn.execute("SELECT 1 FROM seen WHERE id = ?", (item_id,))
    return cur.fetchone() is not None


def mark_seen(conn: sqlite3.Connection, item_id: str, it: Item) -> None:
    conn.execute(
        "INSERT OR IGNORE INTO seen (id, feed, title, link, published, first_seen) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        (
            item_id,
            it.feed,
            it.title,
            it.link,
            it.published.isoformat() if it.published else "",
            datetime.now(timezone.utc).isoformat(),
        ),
    )
    conn.commit()


# --------------------------------------------------------------------------
# Fetching
# --------------------------------------------------------------------------
def fetch_feed(url: str, timeout: int = 15):
    """feedparser can be handed a URL directly, but doing our own request
    lets us set a timeout and a real User-Agent (some sites block the
    default urllib agent)."""
    if url.startswith("file://"):
        return feedparser.parse(url)
    req = urllib.request.Request(
        url, headers={"User-Agent": "Mozilla/5.0 (compatible; CyberNewsMonitor/1.0)"}
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        raw = resp.read()
    return feedparser.parse(raw)


def load_feeds(feeds_file: Optional[str]) -> list:
    if not feeds_file:
        return DEFAULT_FEEDS
    with open(feeds_file, "r", encoding="utf-8") as f:
        feeds = json.load(f)
    if not isinstance(feeds, list) or not all("url" in f for f in feeds):
        sys.exit("Feeds file must be a JSON list of {\"name\": ..., \"url\": ...} objects.")
    return feeds


# --------------------------------------------------------------------------
# Core run
# --------------------------------------------------------------------------
def run_once(args, conn: sqlite3.Connection, feeds: list) -> list:
    now = datetime.now(timezone.utc)
    lookback_seconds = args.lookback_hours * 3600
    new_items: list[Item] = []

    for feed in feeds:
        name, url = feed["name"], feed["url"]
        try:
            parsed = fetch_feed(url)
        except Exception as exc:  # network hiccups shouldn't kill the whole run
            print(f"  [warn] could not fetch {name} ({url}): {exc}", file=sys.stderr)
            continue

        if getattr(parsed, "bozo", 0) and not parsed.entries:
            print(f"  [warn] {name}: feed did not parse cleanly and had no entries", file=sys.stderr)
            continue

        for entry in parsed.entries:
            iid = entry_id(entry)
            if has_seen(conn, iid):
                continue

            title = entry.get("title", "(no title)")
            link = entry.get("link", "")
            summary = entry.get("summary", "") or entry.get("description", "")
            published = parse_published(entry)

            it = Item(feed=name, title=title, link=link, summary=summary, published=published)
            mark_seen(conn, iid, it)  # mark seen regardless, so we never re-process it

            if published and (now - published).total_seconds() > lookback_seconds:
                continue  # too old to bother surfacing, but it's now recorded as seen

            categories, cves = categorize(title, summary)
            it.categories, it.cves = categories, cves

            if categories or args.all:
                new_items.append(it)

    new_items.sort(key=lambda it: it.sort_key(), reverse=True)
    return new_items


# --------------------------------------------------------------------------
# Output
# --------------------------------------------------------------------------
def fmt_item(it: Item, color: bool) -> str:
    ts = it.published.strftime("%Y-%m-%d %H:%M UTC") if it.published else "date unknown"
    tags = ", ".join(it.categories) if it.categories else "uncategorized"
    if color:
        tag_str = f"\033[1;33m[{tags}]\033[0m"
        title_str = f"\033[1m{it.title}\033[0m"
    else:
        tag_str = f"[{tags}]"
        title_str = it.title
    lines = [f"{tag_str} {title_str}", f"    {it.feed} · {ts}", f"    {it.link}"]
    if it.cves:
        lines.append(f"    CVEs: {', '.join(it.cves)}")
    return "\n".join(lines)


def print_report(items: list, color: bool) -> None:
    if not items:
        print("No new matching items this run.")
        return
    grouped: dict[str, list] = {}
    for it in items:
        key = ", ".join(it.categories) if it.categories else "Other"
        grouped.setdefault(key, []).append(it)

    print(f"\n{len(items)} new item(s) found:\n")
    for it in items:
        print(fmt_item(it, color))
        print()


def write_digest(items: list, digest_dir: str) -> Optional[str]:
    if not items:
        return None
    out_dir = Path(digest_dir).expanduser()
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d_%H%M")
    out_path = out_dir / f"cyber-digest-{stamp}.md"
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(f"# Cybersecurity digest — {stamp} UTC\n\n")
        for category in CATEGORY_ORDER + ["Other"]:
            bucket = [it for it in items if (category in it.categories) or
                      (category == "Other" and not it.categories)]
            if not bucket:
                continue
            f.write(f"## {category}\n\n")
            for it in bucket:
                ts = it.published.strftime("%Y-%m-%d %H:%M UTC") if it.published else "date unknown"
                f.write(f"- **{it.title}**  \n  {it.feed} · {ts}  \n  {it.link}\n")
                if it.cves:
                    f.write(f"  CVEs: {', '.join(it.cves)}\n")
                f.write("\n")
    return str(out_path)


def get_desktop_path() -> Path:
    """Best-effort, cross-platform Desktop folder detection. Falls back to
    the home directory if no Desktop folder can be found or created."""
    home = Path.home()
    system = platform.system()
    candidates = []

    if system == "Windows":
        # %USERPROFILE%\Desktop, even under OneDrive Known Folder Move, where
        # it's a junction into the OneDrive-managed folder -- so this stays
        # provider-agnostic rather than hardcoding a OneDrive-specific path.
        candidates.append(home / "Desktop")
    elif system == "Darwin":
        candidates.append(home / "Desktop")
    else:  # Linux and other Unix-likes
        try:
            result = subprocess.run(
                ["xdg-user-dir", "DESKTOP"], capture_output=True, text=True, timeout=3
            )
            if result.returncode == 0:
                xdg_path = Path(result.stdout.strip())
                if xdg_path and xdg_path != home:
                    candidates.append(xdg_path)  # honors localized folder names, e.g. "Bureau"
        except (OSError, subprocess.SubprocessError):
            pass
        candidates.append(home / "Desktop")

    for c in candidates:
        if c.exists():
            return c

    target = candidates[0] if candidates else home
    try:
        target.mkdir(parents=True, exist_ok=True)
        return target
    except OSError:
        return home  # last resort: drop the file in the home directory instead


def prepend_to_xlsx(items: list, xlsx_path: Path) -> None:
    """Adds new items as rows at the top of the sheet (just below the header),
    pushing previously saved rows further down -- so the file always reads
    newest-first."""
    if not items:
        return

    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Cyber News"
        ws.append(XLSX_HEADERS)
        for cell in ws[1]:
            cell.font = Font(name="Arial", bold=True)
        ws.freeze_panes = "A2"
        for i, width in enumerate(XLSX_COLUMN_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

    ws.insert_rows(2, amount=len(items))

    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")
    for offset, it in enumerate(items):
        row = 2 + offset
        published_str = it.published.strftime("%Y-%m-%d %H:%M") if it.published else ""
        values = [
            it.title,
            ", ".join(it.categories) if it.categories else "Uncategorized",
            ", ".join(it.cves),
            it.feed,
            published_str,
            it.link,
            now_str,
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=value).font = Font(name="Arial")
        link_cell = ws.cell(row=row, column=6)
        link_cell.hyperlink = it.link
        link_cell.font = Font(name="Arial", color="0563C1", underline="single")

    wb.save(xlsx_path)


# --------------------------------------------------------------------------
# HTML report -- contains the same rows as the .xlsx log, rendered as a
# standalone webpage with clickable links. The current row set is embedded
# in the page itself as a JSON blob, so re-runs can load it back, prepend
# the new items, and rewrite the page -- no separate data file needed.
# --------------------------------------------------------------------------
HTML_DATA_RE = re.compile(
    r'<script type="application/json" id="row-data">(.*?)</script>', re.DOTALL
)

CATEGORY_BADGE_CLASSES = {
    "CVE": "badge-cve",
    "Zero-Day/Actively Exploited": "badge-zeroday",
    "Ransomware": "badge-ransomware",
    "Breach": "badge-breach",
}


def load_html_rows(html_path: Path) -> list:
    if not html_path.exists():
        return []
    match = HTML_DATA_RE.search(html_path.read_text(encoding="utf-8"))
    if not match:
        return []
    try:
        return json.loads(match.group(1))
    except json.JSONDecodeError:
        return []


def item_to_row(it: Item, now_str: str) -> dict:
    return {
        "title": it.title,
        "categories": ", ".join(it.categories) if it.categories else "Uncategorized",
        "cves": ", ".join(it.cves),
        "feed": it.feed,
        "published": it.published.strftime("%Y-%m-%d %H:%M") if it.published else "",
        "link": it.link,
        "first_seen": now_str,
    }


def render_category_badges(categories: str) -> str:
    if not categories:
        return ""
    spans = []
    for cat in categories.split(", "):
        css_class = CATEGORY_BADGE_CLASSES.get(cat, "badge-other")
        spans.append(f'<span class="badge {css_class}">{html.escape(cat)}</span>')
    return "".join(spans)


def render_html_report(rows: list) -> str:
    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    table_rows = []
    for row in rows:
        link = row.get("link", "")
        link_cell = (
            f'<a href="{html.escape(link, quote=True)}" target="_blank" rel="noopener noreferrer">'
            f'{html.escape(link)}</a>' if link else ""
        )
        table_rows.append(
            "<tr>"
            f'<td>{html.escape(row.get("title", ""))}</td>'
            f'<td>{render_category_badges(row.get("categories", ""))}</td>'
            f'<td>{html.escape(row.get("cves", ""))}</td>'
            f'<td>{html.escape(row.get("feed", ""))}</td>'
            f'<td>{html.escape(row.get("published", ""))}</td>'
            f'<td>{link_cell}</td>'
            f'<td>{html.escape(row.get("first_seen", ""))}</td>'
            "</tr>"
        )

    header_cells = "".join(f"<th>{html.escape(h)}</th>" for h in HTML_HEADERS)
    # Escape "</" so a title/summary containing "</script>" can't break out of
    # the data blob early -- "\/" is a valid JSON escape for "/", so this is
    # transparent to json.loads() on the next run.
    data_blob = json.dumps(rows).replace("</", "<\\/")

    return f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Cyber Security News</title>
<style>
  :root {{
    --bg: #f7f8fa; --surface: #ffffff; --text: #1a1d21; --muted: #5b6270;
    --border: #e2e5ea; --link: #0b57d0; --header-bg: #14161a; --header-text: #ffffff;
  }}
  @media (prefers-color-scheme: dark) {{
    :root {{
      --bg: #14161a; --surface: #1c1f24; --text: #e8eaed; --muted: #9aa1ac;
      --border: #2b2f36; --link: #8ab4f8; --header-bg: #0e0f12; --header-text: #e8eaed;
    }}
  }}
  * {{ box-sizing: border-box; }}
  body {{
    margin: 0; padding: 2rem; background: var(--bg); color: var(--text);
    font-family: -apple-system, Segoe UI, Roboto, Arial, sans-serif;
  }}
  h1 {{ margin: 0 0 0.25rem; font-size: 1.4rem; }}
  .meta {{ color: var(--muted); margin: 0 0 1.25rem; font-size: 0.9rem; }}
  .table-wrap {{
    overflow-x: auto; border: 1px solid var(--border); border-radius: 8px;
    background: var(--surface);
  }}
  table {{ border-collapse: collapse; width: 100%; min-width: 900px; }}
  th, td {{
    text-align: left; padding: 0.6rem 0.75rem; border-bottom: 1px solid var(--border);
    vertical-align: top; font-size: 0.9rem;
  }}
  thead th {{
    background: var(--header-bg); color: var(--header-text); position: sticky; top: 0;
  }}
  tbody tr:hover {{ background: color-mix(in srgb, var(--text) 4%, transparent); }}
  a {{ color: var(--link); text-decoration: none; word-break: break-all; }}
  a:hover {{ text-decoration: underline; }}
  .badge {{
    display: inline-block; padding: 0.1rem 0.5rem; border-radius: 999px;
    font-size: 0.75rem; font-weight: 600; margin: 0 0.25rem 0.25rem 0; white-space: nowrap;
  }}
  .badge-cve {{ background: #e3edff; color: #0b3d91; }}
  .badge-zeroday {{ background: #fde3e3; color: #8a1c1c; }}
  .badge-ransomware {{ background: #f1e3ff; color: #5a1c8a; }}
  .badge-breach {{ background: #ffe9d1; color: #8a4a0a; }}
  .badge-other {{ background: var(--border); color: var(--muted); }}
  @media (prefers-color-scheme: dark) {{
    .badge-cve {{ background: #1c2b4a; color: #a9c3ff; }}
    .badge-zeroday {{ background: #3a1c1c; color: #ffb3b3; }}
    .badge-ransomware {{ background: #2f1c3a; color: #dcb3ff; }}
    .badge-breach {{ background: #3a2a10; color: #ffcf94; }}
  }}
</style>
</head>
<body>
  <h1>Cyber Security News</h1>
  <p class="meta">{len(rows)} item(s) &middot; last updated {generated}</p>
  <div class="table-wrap">
    <table>
      <thead><tr>{header_cells}</tr></thead>
      <tbody>
        {"".join(table_rows)}
      </tbody>
    </table>
  </div>
  <script type="application/json" id="row-data">{data_blob}</script>
</body>
</html>
"""


def prepend_to_html(items: list, html_path: Path) -> None:
    """Adds new items as rows at the top of the page, ahead of previously
    saved rows -- so the file always reads newest-first, mirroring the
    .xlsx log."""
    if not items:
        return

    existing_rows = load_html_rows(html_path)
    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")
    new_rows = [item_to_row(it, now_str) for it in items]
    combined_rows = new_rows + existing_rows

    html_path.parent.mkdir(parents=True, exist_ok=True)
    html_path.write_text(render_html_report(combined_rows), encoding="utf-8")


def maybe_notify_slack(items: list, webhook_url: Optional[str]) -> None:
    if not webhook_url or not items:
        return
    lines = [f"*{len(items)} new cybersecurity item(s):*"]
    for it in items[:20]:  # keep it sane for chat
        tags = ", ".join(it.categories) if it.categories else "uncategorized"
        lines.append(f"• [{tags}] <{it.link}|{it.title}> — {it.feed}")
    payload = json.dumps({"text": "\n".join(lines)}).encode("utf-8")
    req = urllib.request.Request(
        webhook_url, data=payload, headers={"Content-Type": "application/json"}
    )
    try:
        urllib.request.urlopen(req, timeout=10)
    except Exception as exc:
        print(f"  [warn] Slack notification failed: {exc}", file=sys.stderr)


# --------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------
def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--feeds-file", help="JSON file with a custom feed list (overrides defaults)")
    p.add_argument("--db", default=str(Path.home() / ".cyber_monitor" / "seen.sqlite3"),
                    help="Path to the SQLite dedup database")
    p.add_argument("--all", action="store_true",
                    help="Show every new item, not just CVE/breach/ransomware/zero-day matches")
    p.add_argument("--lookback-hours", type=int, default=48,
                    help="Ignore items older than this many hours (default: 48)")
    p.add_argument("--digest-dir", help="If set, also write a Markdown digest file into this directory")
    p.add_argument("--xlsx-path", help="Override the .xlsx output location (default: cyber_security_news.xlsx on the Desktop)")
    p.add_argument("--no-xlsx", action="store_true", help="Disable writing/updating the .xlsx log on the Desktop")
    p.add_argument(
        "--html-path", nargs="?", const=HTML_DESKTOP_DEFAULT, default=None,
        help="Also write/update an HTML report with clickable links and open it in your "
             "default browser after saving. Optionally pass a path to override the default "
             f"location ({HTML_DEFAULT_FILENAME} on the Desktop)."
    )
    p.add_argument("--slack-webhook", default=os.environ.get("CYBER_MONITOR_SLACK_WEBHOOK"),
                    help="Slack incoming-webhook URL to post new items to "
                         "(or set CYBER_MONITOR_SLACK_WEBHOOK env var)")
    p.add_argument("--no-color", action="store_true", help="Disable ANSI color in console output")
    p.add_argument("--watch", action="store_true", help="Keep running, polling on a fixed interval")
    p.add_argument("--interval", type=int, default=30, help="Minutes between polls in --watch mode (default: 30)")
    return p


def main() -> None:
    args = build_arg_parser().parse_args()
    feeds = load_feeds(args.feeds_file)
    conn = open_db(Path(args.db))
    color = sys.stdout.isatty() and not args.no_color

    def cycle():
        print(f"=== Polling {len(feeds)} feed(s) at {datetime.now(timezone.utc).isoformat(timespec='seconds')} ===")
        items = run_once(args, conn, feeds)
        print_report(items, color)
        if args.digest_dir:
            path = write_digest(items, args.digest_dir)
            if path:
                print(f"Digest written to {path}")
        if not args.no_xlsx:
            xlsx_path = Path(args.xlsx_path).expanduser() if args.xlsx_path else get_desktop_path() / XLSX_DEFAULT_FILENAME
            try:
                prepend_to_xlsx(items, xlsx_path)
                if items:
                    print(f"Added {len(items)} row(s) to the top of {xlsx_path}")
            except PermissionError:
                print(f"  [warn] Could not write {xlsx_path} -- is it open in Excel? Close it and rerun.", file=sys.stderr)
            except Exception as exc:
                print(f"  [warn] Failed to update {xlsx_path}: {exc}", file=sys.stderr)
        if args.html_path:
            if args.html_path == HTML_DESKTOP_DEFAULT:
                html_path = get_desktop_path() / HTML_DEFAULT_FILENAME
            else:
                html_path = Path(args.html_path).expanduser()
            try:
                prepend_to_html(items, html_path)
                if items:
                    print(f"Added {len(items)} row(s) to the top of {html_path}")
                    if not webbrowser.open(html_path.resolve().as_uri()):
                        print(f"  [warn] Could not open a browser automatically -- open {html_path} manually.", file=sys.stderr)
            except OSError as exc:
                print(f"  [warn] Failed to update {html_path}: {exc}", file=sys.stderr)
        maybe_notify_slack(items, args.slack_webhook)

    if not args.watch:
        cycle()
        return

    try:
        while True:
            cycle()
            print(f"Sleeping {args.interval} minute(s)...\n")
            time.sleep(args.interval * 60)
    except KeyboardInterrupt:
        print("\nStopped.")


if __name__ == "__main__":
    main()